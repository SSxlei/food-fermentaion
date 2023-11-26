import os
from openpyxl import load_workbook

def write_bed_file(path, data):
    with open(path, 'w') as bed_file:
        for row in data:
            bed_file.write('\t'.join(map(str, row)) + '\n')

def extract_bed_data(excel_path):
    data = []
    wb = load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        chrom = row[1]  # Second column
        chromStart = row[4]  # Fifth column
        chromEnd = row[5]  # Sixth column

        data.append([chrom, chromStart, chromEnd])

    return data

def read_id_file(id_file_path):
    with open(id_file_path, 'r') as id_file:
        return [line.strip() for line in id_file.readlines()]

def main():
    root = "/mnt/hpc/home/wuq8022600160/xulei/metagenome/fermented_food/bins1/MAG/BGC_up_down"
    id_file_path = "sample_id.txt"  # Specify the correct path to your ID file

    # Read specified_ids from the file
    specified_ids = read_id_file(id_file_path)

    for specified_id in specified_ids:
        excel_path = os.path.join(root, f"{specified_id}.xlsx")
        if os.path.exists(excel_path):
            bed_data = extract_bed_data(excel_path)
            bed_file_path = os.path.join(root, f"{specified_id}.bed")
            write_bed_file(bed_file_path, bed_data)
            print(f"Bed file {bed_file_path} created for {specified_id}")
        else:
            print(f"Excel file not found for {specified_id}")

if __name__ == '__main__':
    main()
