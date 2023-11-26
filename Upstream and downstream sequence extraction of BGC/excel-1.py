import os
import openpyxl

def subtract_one(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    # Select the active sheet
    sheet = wb.active

    # Iterate through rows starting from the second row (assuming the first row is a header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=6):
        for cell in row:
            # Check if the cell value is numeric (try converting to float)
            try:
                numeric_value = float(cell.value)
                # Subtract 1 from the numeric value
                cell.value = numeric_value - 1
            except (ValueError, TypeError):
                # If conversion to float fails, leave the value unchanged
                pass

    # Save the modified workbook
    wb.save(file_path)

def process_excel_files_in_folder(folder_path):
    # Iterate through all files in the folder
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            subtract_one(file_path)

# Example usage
folder_path = '/mnt/hpc/home/wuq8022600160/xulei/metagenome/fermented_food/bins1/MAG/BGC_up_down'
process_excel_files_in_folder(folder_path)
